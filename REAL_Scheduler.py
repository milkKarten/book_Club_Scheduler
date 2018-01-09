#! /usr/bin/env python
'''Created by Seth Karten, 2018'''
'''
Creates REAL groups based on timeslots and REAL leader info
IMPORTANT: Include only firstName, lastName, REAL leader,
		   and availability fields in excel spreadsheet
'''
import openpyxl
import random
import math
import time

class schedule:
	'''Assign people to group with most common mode that works'''
	def assignGroups(self):
		for member in range(4, self.numEntries + 1): # Corresponds to row on excel self.sheet
			firstName = self.sheet.cell(row = member, column = 1).value
			lastName = self.sheet.cell(row = member, column = 2).value
			name = str(firstName) + ' ' + str(lastName)
			if firstName == self.blankCell or name in self.finishedMembers:
				continue
			for currentLeader in range(len(self.groupList) - 1, -1, -1):
				if firstName == self.blankCell or member in self.finishedMembers:
					continue
				isAvailable = self.sheet.cell(row = member, column = self.groupList[currentLeader][2])
				if isAvailable.value == 'Available':
					#print(self.groupList)
					self.groupList[currentLeader][0].append(name)
					self.groupList[currentLeader][1].append(member)
					#exec('group%d[0].append(firstName + ' ' + lastName)'%currentLeader)
					self.finishedMembers.append(name)
					break
			if name not in self.finishedMembers:
				self.membersNotAssigned = True
				if len(self.unassignedMembers) == 0:
					self.unassignedMembers = [[name], [member]]
				else:
					self.unassignedMembers[0].append(name)
					self.unassignedMembers[1].append(member)
				print('(' + str(member) + '): ' + str(firstName) + ' ' + str(lastName) + ' did not join group')

	'''Order leaders by mode using insertion sort'''
	def orderByMode(self):
		orderedGroupList = [self.groupList[0]]
		for k in range(1, len(self.groupList)):
			compare = self.groupList[k][3]
			compareTerm = self.groupList[k]
			for j in range(0, len(orderedGroupList)):
				if compare > orderedGroupList[j][3]:
					#print(str(compare) + '\t' + '>' + '\t' + str(orderedGroupList[j][3]))
					compareTermTemp = orderedGroupList[j]
					compare = orderedGroupList[j][3]
					orderedGroupList[j] = compareTerm
					compareTerm = compareTermTemp
				if j == len(orderedGroupList) - 1:
					#print(str(compare) + '\t' + '<' + '\t' + str(orderedGroupList[j][3]))
					orderedGroupList.append(compareTerm)
					break
		return orderedGroupList

	'''Get rid of duplicates and mark down REAL leaders'''
	def delDupsFindLeads(self):
		for person in range(3, self.numEntries + 1):
			duplicate = False
			firstName = self.sheet[self.firstNameColumn + str(person)].value
			lastName = self.sheet[self.lastNameColumn + str(person)].value
			for possibleDuplicate in range(person + 1, self.numEntries + 1):
				curFirstName = self.sheet[self.firstNameColumn + str(possibleDuplicate)].value
				curLastName = self.sheet[self.lastNameColumn + str(possibleDuplicate)].value
				if firstName == curFirstName and lastName == curLastName:
					self.sheet[self.firstNameColumn + str(person)] = self.blankCell
					self.sheet[self.lastNameColumn + str(person)] = self.blankCell
					duplicate = True
			if not duplicate:
				# Manage number of leaders and create groups for
				# leaders with leaders as only member so far
				if self.sheet[self.isLeaderColumn + str(person)].value == 'Yes':
					self.numLeaders += 1
					groupName = 'Group: ' + str(firstName) + ' ' + str(lastName) + ' -'
					self.realLeaders.append(str(firstName) + ' ' + str(lastName))
					self.realLeaderIndicie.append(person)
					self.groupList.append([[groupName],[person],1, 1]) # names, row nums of names, meeting time column, mode
					#exec('group%d = [[member],[person],[0]]'%self.numLeaders) # name, row num, meeting time column
		#print('Number of Leaders: ' + str(self.numLeaders))
		#print(self.groupList)

	'''Sort Availabilities'''
	def sortAvailabilities(self):
		self.modes = [] # Ordered list of self.modes from largest value to smallest; [numAvailable, columnNum]
		for columnNum in range(4, self.maxColumns + 1):
			numAvailable = 0
			for rowNum in range(4, self.numEntries + 1):
				isAvailable = self.sheet.cell(row = rowNum, column = columnNum)
				if isAvailable.value == 'Available':
					numAvailable += 1
			for i in range(0, len(self.modes) + 1, 1):
				if i == len(self.modes):
					self.modes.append([numAvailable, columnNum])
				elif self.modes[i][0] < numAvailable:
					tempMode = self.modes[i][0]
					tempColumnNum = self.modes[i][1]
					self.modes[i] = [numAvailable, columnNum]
					numAvailable = tempMode
					columnNum = tempColumnNum
			if len(self.modes) == 0:
				self.modes.append([numAvailable, columnNum])
		#print('Modes: ' + str(self.modes))
		#print(len(self.modes))

	'''Assign Leaders to mode meeting times'''
	def assignToModes(self):
		self.finishedLeaders = []
		for modeIndicie in range(self.numLeaders - 1, -1, -1):
			vertice = 0
			for currentLeader in self.realLeaders:
				if currentLeader in self.finishedLeaders:
					#print(str(currentLeader) + ' is done')
					vertice += 1
					continue
				columnIndicie = self.modes[modeIndicie][1]
				isAvailable = self.sheet.cell(row = self.groupList[vertice][1][0], column = columnIndicie)
				if isAvailable.value == 'Available':
					self.groupList[vertice][3] = self.modes[modeIndicie][0]
					self.groupList[vertice][2] = columnIndicie
					self.groupList[vertice][0] = [self.groupList[vertice][0][0] + ' ' + str(self.sheet.cell(row = 3, column = columnIndicie).value)]
					#print(self.groupList[vertice][0])
					#print(vertice)
					#print(currentLeader)
					self.finishedLeaders.append(currentLeader)
					break
				vertice += 1
		# In case leaders need less popular times
		#print(self.realLeaders)
		#print(self.finishedLeaders)
		if len(self.finishedLeaders) != self.numLeaders:
			#print('Less popular times needed...')
			for modeIndicie in range(self.numLeaders, len(self.modes)):
				vertice = 0
				if len(self.finishedLeaders) == self.numLeaders:
					break
				for currentLeader in self.realLeaders:
					if currentLeader in self.finishedLeaders:
						vertice += 1
						continue
					columnIndicie = self.modes[modeIndicie][1]
					isAvailable = self.sheet.cell(row = self.groupList[vertice][1][0], column = columnIndicie)
					#exec('isAvailable = self.sheet.cell(row = group%d[1][vertice], column = self.modes[modeIndicie][1])'%currentLeader)
					if isAvailable.value == 'Available':
						self.groupList[vertice][3] = self.modes[modeIndicie][0]
						self.groupList[vertice][2] = columnIndicie
						self.groupList[vertice][0] = [self.groupList[vertice][0][0] + ' ' + str(self.sheet.cell(row = 3, column = columnIndicie).value)]
						#exec('group%d[2][vertice] = self.modes[modeIndicie][1]'%currentLeader)
						self.finishedLeaders.append(currentLeader)
						#(self.groupList[vertice][0])
						break
					vertice += 1
		#Leaders cannot make any times... time to start switching
		#print(self.groupList)
		if len(self.finishedLeaders) != self.numLeaders or len(self.groupList) != self.numLeaders:
			print('not all leaders have groups')
			for leader in range(0, len(self.realLeaders)):
				if self.realLeaders[leader] not in self.finishedLeaders:
					firstName = self.sheet.cell(row = self.realLeaderIndicie[leader], column = 1).value
					lastName = self.sheet.cell(row = self.realLeaderIndicie[leader], column = 2).value
					print('(' + str(leader) + '): ' + str(firstName) + ' ' + str(lastName) + ' did not create group')
		#print(self.finishedLeaders)

	'''Switch leaders to new timeslot if a member does not fit in a group'''
	def switchLeaders(self):
		orderedModes = [[self.modes[0][0]], [self.modes[0][1]]]
		for i in range(1, len(self.modes)):
			orderedModes[0].append(self.modes[i][0]) # Mode
			orderedModes[1].append(self.modes[i][1]) # timeslot column
		while len(self.unassignedMembers) > 0:
			newMembersToAssign = []
			justAssignedMembers = []
			if self.membersNotAssigned:
				for k in range(0, len(self.unassignedMembers[0])):
					movedMember = False
					allMembersCanMove = False
					for mode in range(0, len(orderedModes[1])):
						inList = False
						for j in range(0, len(self.groupList)):
							if orderedModes[1][mode] == self.groupList[j][2]:
								inList = True
								break
						if inList:
							continue
						member = self.unassignedMembers[1][k]
						isAvailable = self.sheet.cell(row = member, column = orderedModes[1][mode])
						#print(str(isAvailable.value) + ' member: ' + str(member) + ' column: ' + str(orderedModes[1][mode]))
						if isAvailable.value == 'Available':
							for group in range(0, len(self.groupList)):
								allMembersCanMove = True
								for rowNum in self.groupList[group][1]:
									isMemberAvailable = self.sheet.cell(row = rowNum, column = orderedModes[1][mode])
									if isMemberAvailable != 'Available':
										allMembersCanMove = False
										break
								if allMembersCanMove:
									print('WOWWWWWW')
									print('Moving Group')
									self.groupList[group][2] = orderedModes[1][mode]
									self.groupList[group][3] = orderedModes[0][mode]
									self.groupList[group][0].append(self.unassignedMembers[0][k])
									self.groupList[group][1].append(self.unassignedMembers[1][k])
									self.unassignedMembers.remove(self.unassignedMembers[0][k])
									self.unassignedMembers.remove(self.unassignedMembers[1][k])
									substringIndex = self.groupList[group][0][0].index('-') + 1
									subString = self.groupList[group][0][0][:substringIndex]
									self.groupList[group][0][0] = subString + ' ' + str(self.sheet.cell(row = 3, column = self.groupList[group][2]).value)
							if not allMembersCanMove:
								cannotMove = []
								canMove = False
								while True:
									group = random.randint(0, len(self.groupList) - 1)
									if group not in cannotMove:
										isAvailable = self.sheet.cell(row = self.groupList[group][1][0], column = orderedModes[1][mode])
										if isAvailable.value == 'Available':
											canMove = True
											break
										else:
											canMove = False
											cannotMove.append(group)
									if len(cannotMove) == len(self.groupList) - self.numLeaders:
										print('Cannot switch group: ' + str(self.sheet.cell(row = member, column = 1).value) + ' ' + str(self.sheet.cell(row = member, column = 2).value) + ' must pick from an available timeslot manually')
										break
								if canMove:
									print('Moving')
									movedMember = True
									if len(newMembersToAssign) == 0:
										newMembersToAssign = [self.groupList[group][0][1:], self.groupList[group][1][1:]]
									else:
										for mem in range(1, len(self.groupList[group][0])):
											newMembersToAssign[0].append(self.groupList[group][0][mem])
											newMembersToAssign[1].append(self.groupList[group][1][mem])
									self.groupList[group][0] = [self.groupList[group][0][0], self.unassignedMembers[0][k]]
									self.groupList[group][1] = [self.groupList[group][1][0], self.unassignedMembers[1][k]]
									self.groupList[group][2] = orderedModes[1][mode]
									self.groupList[group][3] = orderedModes[0][mode]
									substringIndex = self.groupList[group][0][0].index('-') + 1
									subString = self.groupList[group][0][0][:substringIndex]
									self.groupList[group][0][0] = subString + ' ' + str(self.sheet.cell(row = 3, column = self.groupList[group][2]).value)
						else:
							continue
						if movedMember or allMembersCanMove:
							if len(justAssignedMembers) == 0:
								justAssignedMembers = [[self.unassignedMembers[0][k]], [self.unassignedMembers[1][k]]]
							else:
								justAssignedMembers[0].append(self.unassignedMembers[0][k])
								justAssignedMembers[1].append(self.unassignedMembers[1][k])
							break
				#Assign newMembersToAssign to groups
				newFinishedMembers = []
				tempUnassignedMembers = self.unassignedMembers
				self.unassignedMembers = []

				for member in range(0, len(tempUnassignedMembers[0])):
					try:
						if tempUnassignedMembers[0][member] not in justAssignedMembers[0]:
							if len(newMembersToAssign) == 0:
								newMembersToAssign = [[tempUnassignedMembers[0][member]], [tempUnassignedMembers[1][member]]]
							else:
								newMembersToAssign[0].append(tempUnassignedMembers[0][member])
								newMembersToAssign[1].append(tempUnassignedMembers[1][member])
					except:
						pass
						if len(newMembersToAssign) == 0:
							newMembersToAssign = [[tempUnassignedMembers[0][member]], [tempUnassignedMembers[1][member]]]
						else:
							newMembersToAssign[0].append(tempUnassignedMembers[0][member])
							newMembersToAssign[1].append(tempUnassignedMembers[1][member])

				if len(newMembersToAssign) > 0:
					for member in range(0, len(newMembersToAssign[1])):
						name = newMembersToAssign[0][member]
						if name in newFinishedMembers:
							continue
						for currentLeader in range(0, len(self.groupList)):
							rowNum = newMembersToAssign[1][member]
							columnNum = self.groupList[currentLeader][2]
							isAvailable = self.sheet.cell(row = rowNum, column = columnNum)
							if isAvailable.value == 'Available':
								self.groupList[currentLeader][0].append(name)
								self.groupList[currentLeader][1].append(newMembersToAssign[1][member])
								newFinishedMembers.append(name)
								break
						if name not in newFinishedMembers:
							self.membersNotAssigned = True
							if len(self.unassignedMembers) == 0:
								self.unassignedMembers = [[name], [newMembersToAssign[1][member]]]
							else:
								self.unassignedMembers[0].append(name)
								self.unassignedMembers[1].append(newMembersToAssign[1][member])
							print(str(name) + ' did not join group')
							#print(self.unassignedMembers)

	'''Assign each group average number of group members'''
	def averageNumGroupMembers(self):
		totalNumMembers = len(self.finishedMembers)
		#print(totalNumMembers)
		#print(len(self.groupList))
		membersPerGroup = math.floor(totalNumMembers / len(self.groupList))
		#print(membersPerGroup)
		startingTime = time.time()
		while time.time() - startingTime < (self.timeout * 3.0):
			for group in range(0, len(self.groupList)):
				while (len(self.groupList[group][0]) < membersPerGroup + 1):
					secs = time.time()
					if (secs - startingTime) > self.timeout:
						break
					transferFrom = random.randint(-1, len(self.groupList) - 1)
					if len(self.groupList[transferFrom][0]) > membersPerGroup:
						availSwitchers = [] # names
						availSwitchersIndicie = [] # row index
						for member in range(1, len(self.groupList[transferFrom][1])):
							#print(self.groupList[transferFrom][1][member])
							if self.sheet.cell(row = self.groupList[transferFrom][1][member], column = self.groupList[group][2]).value == 'Available':
								availSwitchers.append(self.groupList[transferFrom][0][member])
								availSwitchersIndicie.append(self.groupList[transferFrom][1][member])

						if len(availSwitchers) > 0:
							transferMember = random.choice(availSwitchers)
							transferIndex = availSwitchers.index(transferMember)
							#print(transferIndex)
							transferMemberRow = availSwitchersIndicie[transferIndex]
							self.groupList[transferFrom][0].remove(transferMember)
							self.groupList[transferFrom][1].remove(transferMemberRow)
							self.groupList[group][0].append(transferMember)
							self.groupList[group][1].append(transferMemberRow)
				while (len(self.groupList[group][0]) > membersPerGroup + 2):
					secs = time.time()
					if (secs - startingTime) > self.timeout:
						break
					transferTo = random.randint(0, len(self.groupList) - 1)
					if transferTo == group:
						continue
					#print('transfer to ' + str(transferTo))
					availSwitchers = [] # names
					availSwitchersIndicie = [] # row index
					for member in range(1, len(self.groupList[group][1])):
						#print(self.groupList[transferTo][1][member])
						#print(str(self.groupList[group][1][member]) + ' ' + str(self.groupList[transferTo][2]))
						if self.sheet.cell(row = self.groupList[group][1][member], column = self.groupList[transferTo][2]).value == 'Available':
							availSwitchers.append(self.groupList[group][0][member])
							availSwitchersIndicie.append(self.groupList[group][1][member])
						if (len(self.groupList[group][0]) - len(availSwitchers)) <= (membersPerGroup + 2):
							break
					if len(availSwitchers) > 0:
						transferMember = random.choice(availSwitchers)
						transferIndex = availSwitchers.index(transferMember)
						#print(transferIndex)
						transferMemberRow = availSwitchersIndicie[transferIndex]
						self.groupList[transferTo][0].append(availSwitchers[transferIndex])
						self.groupList[transferTo][1].append(transferMemberRow)
						self.groupList[group][0].remove(availSwitchers[transferIndex])
						self.groupList[group][1].remove(transferMemberRow)

	'''Save file with updated info'''
	def saveSheet(self):
		if self.sheet['C1'].value == self.blankCell:
			self.wb.create_sheet('Groups')
			self.sheet['C1'].value = 'Clear this cell if Groups self.sheet is deleted'
		else:
			sheetName = self.wb.get_sheet_by_name('Groups')
			self.wb.remove_sheet(sheetName)
			self.wb.create_sheet('Groups')
		groupSheet = self.wb.get_sheet_by_name('Groups')
		for i in range(1, len(self.groupList) + 1):
			cell = str(self.sheet.cell(row = 3, column = i + 3))
			cell = cell[15:]
			if cell[2] == '>':
				cell = cell[0:2]
			groupSheet.cell(row = 1, column = i).value = self.sheet[cell].value
			numMembers = len(self.groupList[i - 1][0])
			#exec('numMembers = len(group%d[0])'%i)
			for j in range(0, numMembers):
				groupSheet.cell(row = j + 1, column = i).value = self.groupList[i - 1][0][j]
				#exec('groupSheet.cell(row = j, column = i).value = group%d[0][j]'%i)
		self.wb.save(self.wbName)

	'''Check Groups'''
	def checkGroups(self):
		for group in self.groupList:
			for member in range(0, len(group[1])):
				isAvailable = self.sheet.cell(row = group[1][member], column = group[2])
				if isAvailable.value != 'Available':
					print(str(group[0][member]) + ' did not join a proper timeslot at ' + str(self.sheet.cell(row = 3, column = group[2]).value))

	def __init__(self):
		self.groupList = [] # names, row nums of names, meeting time column, mode
		self.unassignedMembers = [] # names, row nums of names
		self.membersNotAssigned = False
		self.finishedMembers = []

		'''Set to name and directory of excel file'''
		self.wbName = 'rutgers_engineers_assessing_literature_real_program_availability_form.xlsx'
		self.wb = openpyxl.load_workbook(self.wbName)
		self.sheet = self.wb.get_sheet_by_name('Sheet1')

		self.numEntries = self.sheet.max_row
		self.maxColumns = 45
		self.firstNameColumn = 'A'
		self.lastNameColumn = 'B'
		self.isLeaderColumn = 'C'
		self.blankCell = self.sheet['A1'].value
		self.numLeaders = 0
		self.realLeaders = []
		self.realLeaderIndicie = []
		self.timeout = 0.5 # seconds for checking group lengths

		'''Label top columns by number'''
		for i in range(4, self.sheet.max_column + 1):
			self.sheet.cell(row = 1, column = i).value = i

		self.delDupsFindLeads()
		self.sortAvailabilities()
		self.assignToModes()
		self.groupList = self.orderByMode()
		self.finishedMembers = self.finishedLeaders
		self.assignGroups()
		self.switchLeaders()
		self.averageNumGroupMembers()
		self.checkGroups()
		self.saveSheet()

sched = schedule()
